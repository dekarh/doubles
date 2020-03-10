# Ищем повторяющиеся телефоны и переносим строчки с повторением в другой tab excel
import json
import requests
import openpyxl
import os
import re
from datetime import datetime, timedelta

from lib import l, s, fine_phone, format_phone

ROWS_IN_HEADER = 3

if __name__ == '__main__':
    # Если нет xlsx или их несколько - выходим
    all_files = os.listdir(path=".")
    has_xlsx = 0
    xlsx = ''
    for all_file in all_files:
        if all_file.endswith(".xlsx"):
            xlsx = all_file
            has_xlsx += 1
    if not has_xlsx:
        print('ОШИБКА - Нет ни одного xlsx файла - фильтровать нечего\n '
              'Для фильтрации дублей создайте один xlsx файл, в котором должны быть один столбец с '
              'номерами телефонов и любые другие данные')
    elif has_xlsx > 1:
        print('ОШИБКА - Голова кружится от количества файлов\n '
              'Для фильтрации дублей оставьте только один xlsx файл, в котором должны быть один столбец с '
              'номерами телефонов и любые другие данные')
    else:
        print('Начинаем фильтрацию дублей\n')
        # Создаем файл с отфильтрованными данными
        wb_log = openpyxl.Workbook(write_only=True)
        ws_clear = wb_log.create_sheet('Без дублей')
        ws_doubles = wb_log.create_sheet('Дубли')
        ws_source = wb_log.create_sheet('Исходный')
        # Определяем номера колонок
        wb = openpyxl.load_workbook(filename=xlsx, read_only=True)
        ws = wb[wb.sheetnames[0]]
        phone_column = -1
        double_headers = ['№ строки в исходнике', '№ строки в которой дубль', '№ дублирующегося телефона']
        clear_headers = ['№ строки']
        ok = False
        for i, row in enumerate(ws):
            for j, cell in enumerate(row):
                if i > ROWS_IN_HEADER - 1:
                    line = s(cell.value)
                    if line.find('(') > -1 and line.find(')') > -1:
                        code = line.split('(')[1].split(')')[0]
                        if len(code) == 3 and len(re.sub(r'[^0-9]', '', code)) == 3:
                            phone_column = j
                            ok = True
                            break
                elif i == ROWS_IN_HEADER - 1:
                    double_headers.append(cell.value)
                    clear_headers.append(cell.value)
            if ok:
                break
        ws_clear.append(clear_headers)
        ws_doubles.append(double_headers)
        ws_source.append(clear_headers)
        all_phones = {}
        if phone_column > -1:
            for i, row in enumerate(ws):
                if i > ROWS_IN_HEADER - 1:
                    row_dim = []
                    for cell in row:
                        row_dim.append(cell.value)
                    phones_str = str(row[phone_column].value).strip(' ').strip('\n').strip('\t')
                    phones_str = phones_str.replace('\n\n','\n').replace('\n\n','\n').replace('\n\n','\n')
                    phones = phones_str.split('\n')
                    has_double = False
                    for phone in phones:
                        if format_phone(phone):
                            if format_phone(phone) in all_phones.keys():
                                double_phone = fine_phone(phone)
                                double_line = all_phones[format_phone(phone)]
                                has_double = True
                            else:
                                all_phones[format_phone(phone)] = i
                    if has_double:
                        ws_doubles.append([i, double_line, double_phone] + row_dim)
                    else:
                        ws_clear.append([i] + row_dim)
                    ws_source.append([i] + row_dim)
            if not os.path.exists('cleared'):
                os.makedirs('cleared')
            wb_log.save('cleared/' + datetime.now().strftime("%d-%m-%Y_%H-%M") + '.xlsx')
            print('Фильтрация закончена. Результат в папке cleared')
        else:
            # Если нет какой-нибудь колонки - выходим
            if phone_column == -1:
                print('В файле', xlsx, 'нет столбца с телефонами.')
